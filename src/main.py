import os
import math
import openpyxl as op
import plotly.graph_objects as go
import plotly.subplots as sp

timesheetsLocation = "sumtimesheets/Excel/"
debugCellRead = False
debugHourCount = False
showByDay = True
ActualHrs = 6785.3 - 716.75 # Total hours less holiday entitlement

def main():

    hours = genHourDict()
    hoursByDay = genHoursByDayDicts()
    fileCount = 0

    # Iterate over all files in directory specified
    for file in os.scandir(getTimesheetDirPath()):
        if file.is_file():
            if debugCellRead or debugHourCount:
                print(f"[{fileCount}] | ", end = " ")
                print(file.name)
            fileCount += 1
            sheet = loadSheet(file.path)
            timeCells, readSheetTotal = getTimeCells(sheet)
            
            if debugCellRead:
                printCells(timeCells)
                print()
            countWorkedHours(hours, hoursByDay, timeCells, readSheetTotal)
    countedHours = sumHours(hours)
    if debugHourCount:
        print(f"Total counted hours: {countedHours} | Actual: {ActualHrs} (Error: {round(countedHours - ActualHrs, 1)} | {round((countedHours - ActualHrs) / ActualHrs, 1) * 100}% error)")
        print()
    
    showFigure(hours, hoursByDay, fileCount)

""" Load specific sheet from workbook at provided path. """
def loadSheet(path):
    script_dir = os.path.dirname(__file__) 
    rel_path = path 
    abs_file_path = os.path.join(script_dir, rel_path)
    wb = op.load_workbook(abs_file_path, data_only=True)
    return wb.active

""" Reads relevant cells from provided sheet and returns them in a 2D array. """
def getTimeCells(sheet):
    timeCells = [] # Directly holds cells containing times from timesheet

    # Rows 6 thru 12
    for r in range(6, 13):
        new = [] # Holding var for latest row of cells read in

        # Columns B thru I (regular shifts)
        for c in range(2, 10):
            if sheet.cell(r, c).value != None:
                new.append(sheet.cell(r, c).value)
            else:
                new.append(0) # Read empty cells as zeros

        # Column N (get-out)
        if sheet.cell(r, 14).value != None:
            new.append(sheet.cell(r, 14).value)
        else:
            new.append(0)
        timeCells.append(new)

    return timeCells, round(sheet.cell(14, 7).value, 1)

def printCells(cells):
    for y in range(len(cells)):
        match y:
            case 0:
                print("Sun:", end = ' ')
            case 1:
                print("Mon:", end = ' ')
            case 2:
                print("Tue:", end = ' ')
            case 3:
                print ("Wed:", end = ' ')
            case 4:
                print ("Thu:", end = ' ')
            case 5:
                print ("Fri:", end = ' ')
            case 6:
                print ("Sat:", end = ' ')
        for x in range(len(cells[0])):
            print(cells[y][x], end = ' ')
        print()

""" Takes 2D array containing cells read from timesheet and maps each hour worked to the hours dictionary 12am thru 11pm"""
def countWorkedHours(hours, hoursByDay, cells, readSheetTotal):
    timesheetHours = 0

    for day in range(7):
        if debugHourCount:
            print(f"{indToDay[day]}: ", end = ' ')

        # For each shift start/end time pair...
        for shift in range(0, 8, 2):

            # Check shift has both a start AND end time
            if (cells[day][shift] != 0) and (cells[day][shift + 1] != 0):

                startTime = cells[day][shift].hour + round(cells[day][shift].minute / 60, 2)
                endTime = cells[day][shift + 1].hour + round(cells[day][shift + 1].minute / 60, 2)

                # Account for a shift finishing at midnight (00:00)
                if math.trunc(endTime) == 0:
                    endTime =+ 24

                # For each hour spanned by the shift, add one to relevant hour
                for hr in range(math.trunc(startTime), math.trunc(endTime)):
                    hours[hr] += 1
                    hoursByDay[indToDay[day]][hr] += 1
                    timesheetHours += 1

                if startTime % 1 > 0:
                    hours[hr] -= startTime % 1
                    hoursByDay[indToDay[day]][hr] -= startTime % 1
                    timesheetHours -= startTime % 1
                if endTime % 1 > 0:
                    hours[hr] += endTime % 1
                    hoursByDay[indToDay[day]][hr] += endTime % 1
                    timesheetHours += endTime % 1

                if debugHourCount:
                    print(f"{endTime - startTime}", end = ' ')
        
        # Count get-outs. Takes start time from end of evening/night shift, or else assumes 10pm
        # If there's a get-out at all
        if cells[day][8] != 0:
            if cells[day][8].hour != 0:

                startTime = 22

                # If night shift exists...
                if cells[day][7] != 0:
                    startTime = cells[day][7].hour + round(cells[day][7].minute / 60, 2)

                # Else if evening shift exists...
                elif cells[day][5] != 0:
                    startTime = cells[day][5].hour + round(cells[day][5].minute / 60, 2)

                # If No evening/night shift, assume get-out starts at 10pm
                else:
                    startTime = 22

                endTime = startTime + cells[day][8].hour + round(cells[day][8].minute / 60, 2)

                for hr in range(math.trunc(startTime), math.trunc(endTime)):
                    hours[hr % 24] += 1
                    hoursByDay[indToDay[day]][hr % 24] += 1
                    timesheetHours += 1
                if startTime % 1 > 0:
                    hours[hr % 24] -= startTime % 1
                    hoursByDay[indToDay[day]][hr % 24] -= startTime % 1
                    timesheetHours -= startTime % 1
                if endTime % 1 > 0:
                    hours[hr % 24] += endTime % 1
                    hoursByDay[indToDay[day]][hr % 24] += endTime % 1
                    timesheetHours += endTime % 1

                if debugHourCount:
                        print(f"GO: {cells[day][8].hour + round(cells[day][8].minute / 60, 2)}", end = ' ')

        if debugHourCount:
            print()

    if debugHourCount:
        print(f"Counted: {timesheetHours} | Actual: {readSheetTotal}")

""" Sum and return total hours counted. """
def sumHours(hours):
    return sum([hours[hr] for hr in hours])

def showFigure(hours, hoursByDay, weekCount):

    if not showByDay:
        vals = list(hours.values())
        keys = [str(k).rjust(2, "0") + ":00" for k in hours.keys()]

        fig = go.Figure(
            data=[go.Bar(y=vals, x=keys)],
            layout_title_text = f"Hours worked and when (over {weekCount} weeks)"
        )

        fig.update_xaxes(title_text = "Start time")
        fig.update_yaxes(title_text = "Count")
        fig.show()
    else:

        fig = sp.make_subplots( 
            rows = 2,
            cols = 4,
            subplot_titles=(list(hoursByDay.keys()) + ["All"])
        )

        r = 1
        c = 1
        for day in hoursByDay:
            vals = list(hoursByDay[day].values())
            keys = [str(k).rjust(2, "0") + ":00" for k in hoursByDay[day].keys()]

            fig.add_trace(
                go.Bar(y=vals, x=keys, name=day),
                row= r,
                col= c,
            )
            if c == 4:
                r += 1
                c = 0
            c += 1
        
        vals = list(hours.values())
        keys = [str(k).rjust(2, "0") + ":00" for k in hours.keys()]

        fig.add_trace(
                go.Bar(y=vals, x=keys, name="All"),
                row= 2,
                col= 4
            )
        
        fig.update_yaxes(range=[0, 130])
        fig.update_yaxes(range=[0, 700], row=2, col=4)
        fig.show()

""" Returns the absolute path to the timesheets to process. """
def getTimesheetDirPath():
    script_dir = os.path.dirname(__file__) 
    rel_path = timesheetsLocation
    return os.path.join(script_dir, rel_path)

def genIndToDayDict():
    d = {
        0 : "Sun",
        1 : "Mon",
        2 : "Tue",
        3 : "Wed",
        4 : "Thu",
        5 : "Fri",
        6 : "Sat"
    }
    return d

def genHourDict():
    h = {}
    for i in range(24):
        h[i] = 0
    return h

def genHoursByDayDicts():
    d = {}
    for day in indToDay:
        d[indToDay[day]] = genHourDict()
    return d


indToDay = genIndToDayDict()

main()